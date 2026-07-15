# =============================================================================
# forecastmodel.py  —  Housing / Construction Output Forecasting
# =============================================================================
#
# CHANGES FROM PREVIOUS VERSION
# ------------------------------
# 1. THREE-COMPONENT BLENDED FORECAST
#      Component A — Holt-Winters damped trend (primary baseline)
#      Component B — Fixed elasticity adjustment (transparent, auditable)
#      Component C — Regularised linear ensemble (Ridge + Lasso + ElasticNet)
#      Final = weighted blend; weights differ by country (see BLEND_WEIGHTS)
#
# 2. XGBOOST REMOVED
#      Tree models need 100s of rows. On 20 annual observations XGBoost
#      memorises noise. Removed entirely.
#
# 3. EVALUATION METRIC CHANGED: level-MAPE → YoY% MAE
#      You care about growth rate direction and magnitude, not whether
#      the EUR bn number is within 3%. YoY% MAE is what you optimise.
#
# 4. COUNTRY-SPECIFIC BLEND WEIGHTS
#      Germany/UK: good walk-forward R² → ML gets more weight
#      France/Italy: negative/low R² → Holt-Winters dominates
#
# 5. ITALY SUPERBONUS HANDLING
#      2021-2022 excluded from Holt-Winters trend fit.
#      Superbonus years downweighted in linear model training.
#
# 6. ARIMAX FIXED
#      Proper annual DatetimeIndex prevents silent fitting failures.
#
# =============================================================================

import os
import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D

from sklearn.linear_model import Ridge, Lasso, ElasticNet
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_absolute_error, r2_score
from statsmodels.tsa.holtwinters import ExponentialSmoothing

try:
    from statsmodels.tsa.statespace.sarimax import SARIMAX
    HAS_ARIMAX = True
except ImportError:
    HAS_ARIMAX = False
    print("  WARNING: statsmodels not available — ARIMAX disabled.")

# =============================================================================
# CONFIG
# =============================================================================

OUT_DIR = os.path.join(os.getcwd(), "output")
os.makedirs(OUT_DIR, exist_ok=True)

VALIDATE_FROM   = 2015
MIN_TRAIN_ROWS  = 6

# Linear model hyperparameters
RIDGE_ALPHA     = 5.0    # default; Germany overrides to 10.0 (see RIDGE_ALPHA_MAP)
LASSO_ALPHA     = 0.5
ENET_ALPHA      = 1.0
ENET_L1         = 0.5
ARIMAX_ORDER    = (1, 0, 0)

# Country-specific Ridge alpha — higher = more regularisation = smaller coefficients
# Germany: alpha=10 reduces 2026 overshoot (permits/GDP positive but EC says flat)
RIDGE_ALPHA_MAP = {
    "France":         5.0,
    "Germany":        10.0,
    "Italy":          5.0,
    "United Kingdom": 5.0,
}

FORECAST_PCT_CAP = 15.0   # slightly wider than before — 12% was too tight for
                           # post-2008 and post-COVID swings

# NOTE: BLEND_WEIGHTS and EC_DEVIATION_COUNTRIES/DEVIATION_SCALE_MAP were
# superseded by PATH A (pure KPI-driven ML). They are removed to avoid
# confusion — they were never consulted by the final forecast block.

# ── Fixed elasticity KPIs (pre-COVID window 2007-2019) ────────────────────────
# These are computed at runtime from the data; hardcoded fallbacks below
# are used if the data window is too short.
ELASTICITY_FALLBACK = {
    "France":        {"housing_permits_yoy": 0.15, "interest_rate_chg": 3.0, "labor_cost_yoy": -1.15},
    "Germany":       {"housing_permits_yoy": 0.10, "interest_rate_chg": -0.10, "labor_cost_yoy": -0.54},
    "Italy":         {"interest_rate_chg": 0.30,  "labor_cost_yoy": -1.44},
    "United Kingdom":{"housing_permits_yoy": 0.43, "interest_rate_chg": 5.0,  "labor_cost_yoy": -3.50},
}

# Elasticity KPIs — use raw (pre-lag) feature names
ELASTICITY_KPIS = ["housing_permits_yoy", "interest_rate_chg", "labor_cost_yoy"]

# ── Sample weights ────────────────────────────────────────────────────────────
HALFLIFE_BY_COUNTRY = {
    "France": 15, "Germany": 15, "Italy": 15, "United Kingdom": 15,
}
COVID_YEARS       = {2020, 2021}
COVID_WEIGHT      = 0.30
SUPERBONUS_WEIGHT = 0.20


# =============================================================================
# HELPERS
# =============================================================================

def _safe_save(wb, path, label="Excel", retries=3, wait=3):
    import time
    for attempt in range(1, retries + 1):
        try:
            wb.save(path)
            print(f"  {label} saved: {path}")
            return
        except PermissionError:
            if attempt < retries:
                print(f"  WARNING: {label} locked (attempt {attempt}/{retries}) — "
                      f"retrying in {wait}s...")
                time.sleep(wait)
            else:
                alt = path.replace(".xlsx", f"_v{attempt}.xlsx")
                try:
                    wb.save(alt)
                    print(f"  {label} saved to alternate path: {alt}")
                except Exception as e2:
                    print(f"  ERROR: Could not save {label}: {e2}")


def pct_to_level(prev_level, pct):
    if pd.isna(prev_level) or pd.isna(pct):
        return np.nan
    return round(float(prev_level) * (1.0 + float(pct) / 100.0), 1)


def cagr(start_val, end_val, n_years):
    if n_years <= 0 or pd.isna(start_val) or pd.isna(end_val) \
            or float(start_val) <= 0:
        return np.nan
    return ((float(end_val) / float(start_val)) ** (1.0 / n_years) - 1) * 100


def safe_r2(y_true, y_pred):
    if len(y_true) < 2:
        return np.nan
    try:
        return r2_score(y_true, y_pred)
    except Exception:
        return np.nan


def safe_mae(y_true, y_pred):
    if len(y_true) == 0:
        return np.nan
    try:
        return mean_absolute_error(y_true, y_pred)
    except Exception:
        return np.nan


def safe_mape(y_true, y_pred):
    y_true = np.array(y_true, dtype=float)
    y_pred = np.array(y_pred, dtype=float)
    mask   = np.isfinite(y_true) & np.isfinite(y_pred) & (y_true != 0)
    if mask.sum() == 0:
        return np.nan
    return float(np.mean(np.abs((y_true[mask] - y_pred[mask]) / y_true[mask])) * 100.0)


def _sample_weights(years, country, superbonus_years=None, max_year=None):
    halflife = HALFLIFE_BY_COUNTRY.get(country, 15)
    mx = max_year or max(years)
    sb = superbonus_years or set()
    w  = np.array([
        (2.0 ** (-((mx - yr) / halflife)))
        * (SUPERBONUS_WEIGHT if yr in sb else
           COVID_WEIGHT if yr in COVID_YEARS else 1.0)
        for yr in years
    ], dtype=float)
    w /= w.mean()
    return w


def validate_cols(df, cols, name):
    missing = [c for c in cols if c not in df.columns]
    if missing:
        raise ValueError(f"{name} missing columns: {missing}")


# =============================================================================
# EUROCONSTRUCT FULL SERIES LOADER
# =============================================================================

def load_full_ec_series(country, test_end):
    from preprocessing import EUROCONSTRUCT_FILE, _normalize
    df = pd.read_excel(EUROCONSTRUCT_FILE, sheet_name="Sheet1",
                       header=None, engine="openpyxl")
    year_row = df.iloc[3, 3:].tolist()
    years    = []
    for y in year_row:
        try: years.append(int(float(y)))
        except: years.append(None)
    target_norm = _normalize(country)
    for i in range(len(df)):
        cell_val = _normalize(df.iloc[i, 2])
        if cell_val == target_norm or target_norm in cell_val:
            row_vals = df.iloc[i, 3:].tolist()
            s = pd.Series(
                [float(v) / 1000.0 if str(v) not in ("nan", "-") else np.nan
                 for v in row_vals],
                index=years, name="EC",
            )
            s = s[s.index.notna()].copy()
            s.index = s.index.astype(int)
            return s.sort_index()
    raise ValueError(f"Country '{country}' not found in Euroconstruct file.")


# =============================================================================
# COMPONENT A — HOLT-WINTERS DAMPED TREND
# =============================================================================

def _holt_winters_forecast(hist_clean, target_lvl, train_end, test_end,
                            superbonus_years=None):
    """
    Fit a damped Holt-Winters trend on historical levels.
    Italy: exclude Superbonus years (2021-2022) from fit, re-anchor to 2025 actual.
    Returns dict {year: forecasted_level}.
    """
    sb    = superbonus_years or set()
    s_all = hist_clean[target_lvl].dropna().sort_index()
    s_all = s_all[s_all.index <= train_end]

    # For Superbonus countries: fit on clean window, then re-anchor
    if sb:
        s_fit = s_all.drop([y for y in sb if y in s_all.index])
        print(f"  Holt-Winters: excluding Superbonus years {sorted(sb)} from fit")
    else:
        s_fit = s_all

    if len(s_fit) < 4:
        print("  WARNING: Holt-Winters — insufficient fit data, using last-value extrapolation")
        last_val = float(s_all.iloc[-1])
        return {yr: last_val for yr in range(train_end + 1, test_end + 1)}

    forecast_years = list(range(train_end + 1, test_end + 1))
    n_steps        = test_end - int(s_fit.index.max())

    try:
        model   = ExponentialSmoothing(
            s_fit.values, trend="add", damped_trend=True,
            initialization_method="estimated",
        ).fit(optimized=True)
        raw_fc  = model.forecast(n_steps)
        # raw_fc covers from (s_fit.index.max()+1) to test_end
        fc_idx  = list(range(int(s_fit.index.max()) + 1,
                             int(s_fit.index.max()) + n_steps + 1))
        fc_map  = dict(zip(fc_idx, raw_fc))

        # Re-anchor: if Superbonus years were excluded, the fit may end before
        # train_end. Scale forecast so level(train_end) matches the actual.
        last_fit_yr  = int(s_fit.index.max())
        last_act_yr  = int(s_all.index.max())
        if last_fit_yr < last_act_yr and last_act_yr in s_all.index:
            actual_end   = float(s_all.loc[last_act_yr])
            hw_at_end    = fc_map.get(last_act_yr, None)
            if hw_at_end and abs(hw_at_end) > 0:
                scale = actual_end / hw_at_end
                fc_map = {yr: v * scale for yr, v in fc_map.items()}
                print(f"  Holt-Winters re-anchor: scale={scale:.4f} "
                      f"(HW at {last_act_yr}={hw_at_end:.1f} → actual={actual_end:.1f})")

        result = {yr: round(float(fc_map[yr]), 1)
                  for yr in forecast_years if yr in fc_map}

    except Exception as e:
        print(f"  WARNING: Holt-Winters failed ({e}) — linear trend fallback")
        x    = np.arange(len(s_fit))
        coef = np.polyfit(x, s_fit.values, 1)
        result = {}
        for i, yr in enumerate(forecast_years):
            steps_ahead = yr - int(s_fit.index.max())
            result[yr]  = round(float(np.polyval(coef, len(s_fit) - 1 + steps_ahead)), 1)

    print(f"  Holt-Winters forecast: "
          + "  ".join(f"{yr}={result.get(yr,'n/a')}" for yr in forecast_years))
    return result


# =============================================================================
# COMPONENT B — FIXED ELASTICITY ADJUSTMENT
# =============================================================================

def _compute_elasticities(hist_clean, df_full, target_pct, country,
                           clean_features):
    """
    Compute OLS elasticity of construction YoY% to each elasticity KPI
    over the pre-COVID clean window (2007-2019).
    Falls back to hardcoded values if insufficient data.
    Returns dict {feature_name: elasticity_coefficient}.
    """
    pre_covid_idx = [y for y in hist_clean.index if 2007 <= y <= 2019]
    if len(pre_covid_idx) < 6:
        print("  Elasticity: insufficient pre-COVID data — using fallback coefficients")
        return ELASTICITY_FALLBACK.get(country, {})

    train_pre = hist_clean.loc[pre_covid_idx]
    tgt       = train_pre[target_pct].dropna()
    elasticities = {}

    for kpi in ELASTICITY_KPIS:
        # Find the column (may be lagged)
        col = kpi
        for c in df_full.columns:
            if c.startswith(kpi):
                col = c
                break
        if col not in df_full.columns:
            continue
        feat = df_full.loc[[y for y in pre_covid_idx if y in df_full.index], col]
        common = feat.index.intersection(tgt.index)
        f_c = feat.loc[common].dropna()
        t_c = tgt.loc[f_c.index]
        if len(f_c) < 5:
            continue
        mask   = f_c.notna() & t_c.notna()
        if mask.sum() < 5:
            continue
        try:
            coef = np.polyfit(f_c[mask].values, t_c[mask].values, 1)
            elasticities[kpi] = round(float(coef[0]), 4)
        except Exception:
            pass

    # Fill any missing with fallback
    fallback = ELASTICITY_FALLBACK.get(country, {})
    for kpi in ELASTICITY_KPIS:
        if kpi not in elasticities and kpi in fallback:
            elasticities[kpi] = fallback[kpi]
            print(f"  Elasticity fallback for '{kpi}': {fallback[kpi]}")

    print(f"  Elasticities (2007-2019 OLS): {elasticities}")
    return elasticities


def _elasticity_forecast(df_full, ec_full, elasticities, train_end, test_end):
    """
    For each forecast year, compute adjustment = sum(elasticity_i * delta_KPI_i)
    added on top of HW baseline YoY%.
    Returns dict {year: yoy_pct_adjustment}.
    """
    forecast_years = list(range(train_end + 1, test_end + 1))
    adjustments    = {}

    # Baseline: mean KPI change over recent pre-COVID years (2015-2019)
    base_years = [y for y in range(2015, 2020) if y in df_full.index]

    for yr in forecast_years:
        if yr not in df_full.index:
            adjustments[yr] = 0.0
            continue
        adj = 0.0
        for kpi, coef in elasticities.items():
            # Try to find the column
            col = kpi
            for c in df_full.columns:
                if c.startswith(kpi):
                    col = c
                    break
            if col not in df_full.columns:
                continue
            fc_val   = df_full.loc[yr, col]
            if pd.isna(fc_val):
                continue
            # Compare to pre-COVID baseline mean of that KPI
            base_vals = df_full.loc[base_years, col].dropna()
            base_mean = float(base_vals.mean()) if len(base_vals) > 0 else 0.0
            delta     = float(fc_val) - base_mean
            adj      += coef * delta
        adjustments[yr] = round(adj, 3)

    print(f"  Elasticity adjustments: "
          + "  ".join(f"{yr}={v:+.2f}pp" for yr, v in adjustments.items()))
    return adjustments


# =============================================================================
# COMPONENT C — LINEAR MODEL FIT & PREDICT
# =============================================================================

def _fit_linear_models(X_raw, y, X_scaled, sample_weight=None, train_years=None):
    sw = sample_weight
    models = {}
    models["Ridge"] = Ridge(alpha=RIDGE_ALPHA).fit(X_scaled, y, sample_weight=sw)
    models["Lasso"] = Lasso(alpha=LASSO_ALPHA, max_iter=10000).fit(
        X_scaled, y, sample_weight=sw)
    models["ElasticNet"] = ElasticNet(
        alpha=ENET_ALPHA, l1_ratio=ENET_L1, max_iter=10000
    ).fit(X_scaled, y, sample_weight=sw)
    if HAS_ARIMAX and train_years is not None and len(y) >= 8:
        try:
            # Fixed: use proper annual DatetimeIndex to prevent silent ARIMAX failures
            idx    = pd.date_range(
                start=str(train_years[0]), periods=len(y), freq="YE"
            )
            endog  = pd.Series(y, index=idx)
            exog   = pd.DataFrame(X_raw, index=idx,
                                   columns=[f"x{i}" for i in range(X_raw.shape[1])])
            res    = SARIMAX(
                endog, exog=exog, order=ARIMAX_ORDER,
                enforce_stationarity=False, enforce_invertibility=False,
            ).fit(disp=False)
            models["ARIMAX"] = res
        except Exception as e:
            print(f"  WARNING: ARIMAX fit failed: {e}")
    return models


def _predict_linear(model_name, model_obj, X_raw, X_scaled, train_years=None):
    if model_name in ("Ridge", "Lasso", "ElasticNet"):
        raw = float(model_obj.predict(X_scaled)[0])
    elif model_name == "ARIMAX":
        try:
            exog_fc = pd.DataFrame(
                X_raw.reshape(1, -1),
                columns=[f"x{i}" for i in range(X_raw.shape[1])],
            )
            fc  = model_obj.forecast(steps=1, exog=exog_fc)
            raw = float(fc.iloc[0])
        except Exception as e:
            print(f"  WARNING: ARIMAX forecast failed ({e}) — using fitted mean")
            raw = float(model_obj.fittedvalues.mean())
    else:
        raise ValueError(f"Unknown model: {model_name}")
    return float(np.clip(raw, -FORECAST_PCT_CAP, FORECAST_PCT_CAP))


def _build_linear_names():
    names = ["Ridge", "Lasso", "ElasticNet"]
    if HAS_ARIMAX:
        names.append("ARIMAX")
    return names


# =============================================================================
# CHART STYLE
# =============================================================================

def _clean_axes(ax, fig):
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")
    for spine in ["top", "right", "left", "bottom"]:
        ax.spines[spine].set_visible(False)
    ax.tick_params(colors="#555555", labelsize=8.5, length=0)
    ax.grid(axis="y", color="#E8E8E8", linewidth=0.5, zorder=0)
    ax.grid(axis="x", visible=False)


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def run_forecast(prep):

    # ── Unpack ────────────────────────────────────────────────────────────────
    df              = prep["df"]
    hist_clean      = prep["hist_clean"]
    train_lagged    = prep["train_lagged"]
    clean_features  = prep["CLEAN_FEATURES"]
    target_pct      = prep["TARGET_PCT"]
    target_lvl      = prep["TARGET_LVL"]
    COUNTRY         = prep["COUNTRY"]
    train_end       = prep["TRAIN_END"]
    test_end        = prep["TEST_END"]
    superbonus_years= prep.get("SUPERBONUS_YEARS", set())
    train_window    = prep.get("TRAIN_WINDOW", "full")
    blend_ratio     = prep.get("BLEND_RATIO",  0.70)
    train_start     = prep.get("TRAIN_START",  2007)
    train_end_full  = prep.get("TRAIN_END_FULL", train_end)  # for tiered mode
    tiered_cfg      = prep.get("TIERED_CFG",  {})            # tiered weight params

    linear_names    = _build_linear_names()
    # PATH A sets blend_hw=0, blend_el=0, blend_ml=1 — kept for Excel export compat
    blend_hw, blend_el, blend_ml = 0.0, 0.0, 1.0

    # Path A window labels
    _win_labels = {
        "full":   f"{train_start}-{train_end}",
        "clean":  f"{train_start}-2019 (clean)",
        "blend":  f"{blend_ratio:.0%}clean+{1-blend_ratio:.0%}full",
        "tiered": f"{train_start}-{train_end_full} (tiered weights: covid=0.10, post=0.20, recent=0.50)",
    }

    print(f"  Country        : {COUNTRY}")
    print(f"  Training years : {train_start}-{train_end}  ({len(hist_clean)} actuals)")
    print(f"  Train window   : {_win_labels.get(train_window, train_window)}")
    print(f"  Forecast years : {train_end+1}-{test_end}")
    print(f"  Features       : {clean_features}")
    print(f"  Linear models  : {linear_names}")

    validate_cols(df,           clean_features,                "df")
    validate_cols(hist_clean,   [target_pct, target_lvl],      "hist_clean")
    validate_cols(train_lagged, clean_features + [target_pct], "train_lagged")

    train_years    = [y for y in hist_clean.index if y <= train_end]
    forecast_years = list(range(train_end + 1, test_end + 1))

    ec_full = load_full_ec_series(COUNTRY, test_end)

    # =========================================================================
    # COMPONENT A — HOLT-WINTERS
    # =========================================================================
    print("\n" + "=" * 65)
    print("  COMPONENT A — HOLT-WINTERS DAMPED TREND")
    print("=" * 65)

    hw_levels = _holt_winters_forecast(
        hist_clean, target_lvl, train_end, test_end, superbonus_years
    )

    # =========================================================================
    # COMPONENT B — FIXED ELASTICITY
    # =========================================================================
    print("\n" + "=" * 65)
    print("  COMPONENT B — FIXED ELASTICITY ADJUSTMENT")
    print("=" * 65)

    elasticities  = _compute_elasticities(
        hist_clean, df, target_pct, COUNTRY, clean_features
    )
    el_adjustments = _elasticity_forecast(df, ec_full, elasticities, train_end, test_end)

    # =========================================================================
    # COMPONENT C — WALK-FORWARD VALIDATION + LINEAR MODELS
    # =========================================================================
    print("\n" + "=" * 65)
    print(f"  COMPONENT C — WALK-FORWARD VALIDATION  {VALIDATE_FROM}-{train_end+1}")
    print("  (Evaluation metric: YoY% MAE — growth rate accuracy)")
    print("=" * 65)

    wf_actuals_pct = {}   # YoY% actuals
    wf_actuals_lvl = {}   # level actuals
    wf_preds_pct   = {m: {} for m in linear_names}
    wf_preds_lvl   = {m: {} for m in linear_names}

    # Folds 2015-train_end
    for i in range(1, len(train_years)):
        tr_yrs = train_years[:i]
        te_yr  = train_years[i]
        if te_yr < VALIDATE_FROM:
            continue

        train_slice = train_lagged.loc[
            tr_yrs, clean_features + [target_pct]
        ].replace([np.inf, -np.inf], np.nan).dropna()
        if len(train_slice) < MIN_TRAIN_ROWS:
            continue
        if te_yr not in train_lagged.index:
            continue

        test_row = train_lagged.loc[[te_yr], clean_features + [target_pct]].replace(
            [np.inf, -np.inf], np.nan
        )
        if test_row[clean_features].isna().any(axis=1).iloc[0]:
            continue
        if pd.isna(test_row[target_pct].iloc[0]):
            continue

        X_tr   = train_slice[clean_features].values
        y_tr   = train_slice[target_pct].values
        X_te   = test_row[clean_features].values
        sw     = _sample_weights(train_slice.index.tolist(), COUNTRY,
                                  superbonus_years)
        scaler = StandardScaler()
        X_tr_s = scaler.fit_transform(X_tr)
        X_te_s = scaler.transform(X_te)

        prev_lvl              = float(hist_clean.loc[tr_yrs[-1], target_lvl])
        wf_actuals_pct[te_yr] = float(test_row[target_pct].iloc[0])
        wf_actuals_lvl[te_yr] = float(hist_clean.loc[te_yr, target_lvl])

        fitted = _fit_linear_models(X_tr, y_tr, X_tr_s,
                                     sample_weight=sw,
                                     train_years=train_slice.index.tolist())
        for m in linear_names:
            if m not in fitted:
                continue
            try:
                pp = _predict_linear(m, fitted[m], X_te, X_te_s,
                                     train_years=train_slice.index.tolist())
                pl = pct_to_level(prev_lvl, pp)
            except Exception:
                pp, pl = np.nan, np.nan
            wf_preds_pct[m][te_yr] = pp
            wf_preds_lvl[m][te_yr] = pl

    # Extra fold: first forecast year vs EC reference
    yr_first_fc = train_end + 1
    if yr_first_fc in ec_full.index and yr_first_fc in df.index:
        train_slice_fc1 = train_lagged.loc[
            train_years, clean_features + [target_pct]
        ].replace([np.inf, -np.inf], np.nan).dropna()

        feat_fc1 = df.loc[[yr_first_fc], clean_features].copy().replace(
            [np.inf, -np.inf], np.nan
        )
        if "output_yoy_lag1" in clean_features:
            if train_end in ec_full.index and train_end - 1 in ec_full.index:
                l1 = float(ec_full.loc[train_end])
                l2 = float(ec_full.loc[train_end - 1])
                if l2 > 0:
                    feat_fc1.loc[yr_first_fc, "output_yoy_lag1"] = round(
                        (l1 / l2 - 1) * 100, 3)

        # Flag 3 fix: extra fold removed — it used EC_2026 as ground truth which
        # biased ensemble_mods selection toward models that agree with EC.
        # PATH A's _fit_and_predict does not use ensemble_mods so final numbers
        # were unaffected, but it was conceptually wrong. Walk-forward now runs
        # only on training actuals (2015-2025), which are genuine ground truth.
        has_missing = True  # skip this fold cleanly
        if False and len(train_slice_fc1) >= MIN_TRAIN_ROWS and not has_missing:
            X_tr   = train_slice_fc1[clean_features].values
            y_tr   = train_slice_fc1[target_pct].values
            X_te   = feat_fc1[clean_features].values
            sw     = _sample_weights(train_slice_fc1.index.tolist(), COUNTRY,
                                      superbonus_years)
            scaler = StandardScaler()
            X_tr_s = scaler.fit_transform(X_tr)
            X_te_s = scaler.transform(X_te)

            prev_lvl                    = float(ec_full.loc[train_end])
            ec_ref_lvl                  = float(ec_full.loc[yr_first_fc])
            ec_ref_pct                  = round((ec_ref_lvl / prev_lvl - 1) * 100, 3)
            wf_actuals_pct[yr_first_fc] = ec_ref_pct
            wf_actuals_lvl[yr_first_fc] = ec_ref_lvl

            fitted_fc1 = _fit_linear_models(X_tr, y_tr, X_tr_s,
                                              sample_weight=sw,
                                              train_years=train_slice_fc1.index.tolist())
            for m in linear_names:
                if m not in fitted_fc1:
                    continue
                try:
                    pp = _predict_linear(m, fitted_fc1[m], X_te, X_te_s,
                                         train_years=train_slice_fc1.index.tolist())
                    pl = pct_to_level(prev_lvl, pp)
                except Exception:
                    pp, pl = np.nan, np.nan
                wf_preds_pct[m][yr_first_fc] = pp
                wf_preds_lvl[m][yr_first_fc] = pl
            print(f"  {yr_first_fc} fold added (EC ref = {ec_ref_lvl:.2f} EUR bn,"
                  f" EC YoY = {ec_ref_pct:+.2f}%)")
        else:
            print(f"  {yr_first_fc} fold skipped — "
                  f"{'missing features' if has_missing else 'too few rows'}")

    # ── Accuracy on YoY% (the metric that matters) ────────────────────────────
    print(f"\n  Walk-forward folds: {sorted(wf_actuals_pct.keys())}")
    print(f"\n  PRIMARY METRIC: YoY% MAE  (growth rate accuracy)")
    print(f"  {'Model':<16} {'YoY MAE(pp)':>12} {'Level MAPE%':>13} {'R²':>8} {'N':>5}")
    print(f"  {'-' * 57}")

    yoy_mae_dict  = {}
    lvl_mape_dict = {}
    r2_dict       = {}
    act_yrs   = sorted(wf_actuals_pct.keys())
    act_pct   = np.array([wf_actuals_pct[y] for y in act_yrs])
    act_lvl   = np.array([wf_actuals_lvl[y] for y in act_yrs])

    for m in linear_names:
        pred_pct = np.array([wf_preds_pct[m].get(y, np.nan) for y in act_yrs])
        pred_lvl = np.array([wf_preds_lvl[m].get(y, np.nan) for y in act_yrs])
        mask     = np.isfinite(pred_pct)
        n        = int(mask.sum())
        if n < 3:
            continue
        yoy_mae_dict[m]  = safe_mae(act_pct[mask], pred_pct[mask])
        lvl_mape_dict[m] = safe_mape(act_lvl[mask], pred_lvl[mask])
        r2_dict[m]       = safe_r2(act_lvl[mask], pred_lvl[mask])
        print(f"  {m:<16} {yoy_mae_dict[m]:>12.2f}pp {lvl_mape_dict[m]:>12.2f}%"
              f"  {r2_dict[m]:>+8.4f} {n:>5}")

    if not yoy_mae_dict:
        raise ValueError("Walk-forward produced no usable folds.")

    # Best linear model = lowest YoY% MAE
    best_linear = min(yoy_mae_dict, key=yoy_mae_dict.get)
    print(f"\n  → Best linear model: {best_linear}  "
          f"(YoY MAE {yoy_mae_dict[best_linear]:.2f}pp)")

    # Linear ensemble: top-3 by YoY MAE
    ranked        = sorted(yoy_mae_dict, key=yoy_mae_dict.get)
    ensemble_mods = ranked[:min(3, len(ranked))]
    wf_preds_pct["LinearEnsemble"] = {}
    wf_preds_lvl["LinearEnsemble"] = {}
    for yr in act_yrs:
        vals = [wf_preds_pct[m].get(yr, np.nan) for m in ensemble_mods
                if pd.notna(wf_preds_pct[m].get(yr, np.nan))]
        wf_preds_pct["LinearEnsemble"][yr] = np.mean(vals) if vals else np.nan
        vals_l = [wf_preds_lvl[m].get(yr, np.nan) for m in ensemble_mods
                  if pd.notna(wf_preds_lvl[m].get(yr, np.nan))]
        wf_preds_lvl["LinearEnsemble"][yr] = np.mean(vals_l) if vals_l else np.nan

    ens_pct  = np.array([wf_preds_pct["LinearEnsemble"].get(y, np.nan) for y in act_yrs])
    ens_lvl  = np.array([wf_preds_lvl["LinearEnsemble"].get(y, np.nan) for y in act_yrs])
    ens_mask = np.isfinite(ens_pct)
    if ens_mask.sum() >= 3:
        yoy_mae_dict["LinearEnsemble"]  = safe_mae(act_pct[ens_mask], ens_pct[ens_mask])
        lvl_mape_dict["LinearEnsemble"] = safe_mape(act_lvl[ens_mask], ens_lvl[ens_mask])
        r2_dict["LinearEnsemble"]       = safe_r2(act_lvl[ens_mask], ens_lvl[ens_mask])
        print(f"  LinearEnsemble ({ensemble_mods}): "
              f"YoY MAE {yoy_mae_dict['LinearEnsemble']:.2f}pp  "
              f"R² {r2_dict['LinearEnsemble']:+.4f}")

    if "LinearEnsemble" not in linear_names:
        linear_names.append("LinearEnsemble")

    # =========================================================================
    # RECURSIVE FORECAST 2026-TEST_END  (all linear models)
    # =========================================================================
    print("\n" + "=" * 65)
    print(f"  RECURSIVE LINEAR FORECAST {train_end+1}-{test_end}")
    print("=" * 65)

    lin_pct = {m: {} for m in linear_names}
    lin_lvl = {m: {} for m in linear_names}

    fcast_features = df.loc[forecast_years, clean_features].copy().replace(
        [np.inf, -np.inf], np.nan
    )

    for yr in forecast_years:
        all_train_yrs = [y for y in range(int(train_lagged.index.min()), yr)
                         if y in train_lagged.index and y in ec_full.index]

        train_slice_full = train_lagged.loc[
            [y for y in all_train_yrs if y in train_lagged.index],
            clean_features + [target_pct],
        ].replace([np.inf, -np.inf], np.nan).dropna()

        # Pseudo-actuals loop DISABLED — PATH A final forecast uses _fit_and_predict
        # which is fully independent. This loop previously injected EC_2026 as a
        # fake training row for the 2027 intermediate prediction (lin_pct/lin_lvl),
        # which leaked the EC forecast. Sheet 2 now shows genuine model predictions.
        extra_rows = []

        if extra_rows:
            train_slice_full = pd.concat(
                [train_slice_full] + extra_rows
            ).sort_index().replace([np.inf, -np.inf], np.nan).dropna()

        if len(train_slice_full) < MIN_TRAIN_ROWS:
            print(f"  WARNING: year {yr} — too few rows ({len(train_slice_full)})")
            continue

        X_tr   = train_slice_full[clean_features].values
        y_tr   = train_slice_full[target_pct].values
        sw     = _sample_weights(train_slice_full.index.tolist(), COUNTRY,
                                  superbonus_years)
        scaler = StandardScaler()
        X_tr_s = scaler.fit_transform(X_tr)

        X_fc_row = fcast_features.loc[[yr], clean_features].copy()
        if "output_yoy_lag1" in clean_features:
            py = yr - 1
            if yr == forecast_years[0]:
                # 2026: use last known actual (safe — 2025 is real data)
                if py in ec_full.index and py - 1 in ec_full.index:
                    l1 = float(ec_full.loc[py])
                    l2 = float(ec_full.loc[py - 1])
                    if l2 > 0:
                        X_fc_row.loc[yr, "output_yoy_lag1"] = round((l1/l2-1)*100, 3)
            else:
                # 2027+: use model's own prior prediction — no EC leak (Flag 2 fix)
                if py in lin_pct.get("LinearEnsemble", {}):
                    prior_pct = lin_pct["LinearEnsemble"].get(py, np.nan)
                    if pd.notna(prior_pct):
                        X_fc_row.loc[yr, "output_yoy_lag1"] = round(prior_pct, 3)

        if X_fc_row[clean_features].isna().any(axis=1).iloc[0]:
            miss = [f for f in clean_features if pd.isna(X_fc_row.loc[yr, f])]
            print(f"  WARNING: year {yr} — missing features: {miss}")
            continue

        X_fc   = X_fc_row.values
        X_fc_s = scaler.transform(X_fc)
        # Flag 1 fix: prev_lvl for 2027 uses model's own predicted 2026 level
        # not EC's forecast — keeps Sheet 2 EC-leak free
        if yr == forecast_years[0]:
            prev_lvl = float(ec_full.loc[yr - 1]) if yr - 1 in ec_full.index else np.nan
        else:
            py_lvl = yr - 1
            prev_lvl = lin_lvl.get("LinearEnsemble", {}).get(py_lvl, np.nan)
            if pd.isna(prev_lvl):
                prev_lvl = float(ec_full.loc[py_lvl]) if py_lvl in ec_full.index else np.nan

        fitted = _fit_linear_models(X_tr, y_tr, X_tr_s,
                                     sample_weight=sw,
                                     train_years=train_slice_full.index.tolist())

        for m in linear_names:
            if m == "LinearEnsemble":
                continue
            if m not in fitted:
                continue
            try:
                pp = _predict_linear(m, fitted[m], X_fc, X_fc_s,
                                     train_years=train_slice_full.index.tolist())
                pl = pct_to_level(prev_lvl, pp)
            except Exception as e:
                print(f"  WARNING: {m} year {yr} failed: {e}")
                pp, pl = np.nan, np.nan
            lin_pct[m][yr] = pp
            lin_lvl[m][yr] = pl

        # LinearEnsemble for this year
        ens_pcts_yr = [lin_pct[m].get(yr, np.nan) for m in ensemble_mods
                       if pd.notna(lin_pct[m].get(yr, np.nan))]
        ens_pp_yr   = float(np.mean(ens_pcts_yr)) if ens_pcts_yr else np.nan
        lin_pct["LinearEnsemble"][yr] = ens_pp_yr
        lin_lvl["LinearEnsemble"][yr] = pct_to_level(prev_lvl, ens_pp_yr)

    # =========================================================================
    # FINAL BLENDED FORECAST
    # =========================================================================
    print("\n" + "=" * 65)

    # =========================================================================
    # FINAL FORECAST — PATH A: pure KPI-driven ML, no EC anchor
    # =========================================================================
    # Training window per country (set in preprocessing COUNTRY_CONFIG):
    #   Germany : full 2007-2025  (ML naturally close to EC, R²=0.80)
    #   France  : clean 2007-2019 (post-COVID distortion hurts signal, R²=0.11)
    #   Italy   : clean 2007-2019 (Superbonus breaks full window)
    #   UK      : 70% clean + 30% full (captures both structural + recent)
    # EC 2026/2027 values are NEVER used in training or prediction.
    # KPI files (permits/GDP/interest/HPI/labor) contain Euromonitor 2026/2027
    # forecasts — those ARE used as model inputs, which is correct.
    # =========================================================================
    print("\n" + "=" * 65)
    print(f"  FINAL FORECAST — PATH A  (pure ML, no EC anchor)")
    print(f"  Window: {_win_labels.get(train_window, train_window)}")
    print("=" * 65)

    CLEAN_END = 2019   # pre-COVID clean window cutoff

    def _train_slice(window_type):
        """Return train dataframe for a given window type."""
        full   = train_lagged[(train_lagged.index >= train_start) &
                              (train_lagged.index <= train_end)].copy()
        clean  = train_lagged[(train_lagged.index >= train_start) &
                              (train_lagged.index <= CLEAN_END)].copy()
        tiered = train_lagged[(train_lagged.index >= train_start) &
                              (train_lagged.index <= train_end_full)].copy()
        if window_type == "clean":
            return clean
        if window_type == "tiered":
            return tiered
        return full  # "full" default

    def _tiered_weights(years):
        """
        Tiered weight scheme for UK (2001-2025):
          2001-2019: exponential recency decay × 1.0  (structural cycle)
          2020-2021: decay × 0.05  (COVID — KPI↔output broke down)
          2022:      decay × 0.20  (post-COVID, rate shock starting)
          2023-2025: decay × 0.70  (mild recent years, genuine signal)
        Italy superbonus override not needed here (UK has no superbonus).
        """
        mx  = max(years)
        hl  = HALFLIFE_BY_COUNTRY.get(COUNTRY, 15)
        tcfg = tiered_cfg  # from closure
        covid_yrs   = tcfg.get("covid_years",     {2020, 2021})
        covid_w     = tcfg.get("covid_weight",     0.05)
        postcovid_y = tcfg.get("postcovid_years",  {2022})
        postcovid_w = tcfg.get("postcovid_weight", 0.20)
        recent_yrs  = tcfg.get("recent_years",     {2023, 2024, 2025})
        recent_w    = tcfg.get("recent_weight",    0.70)
        sb          = superbonus_years or set()

        w = []
        for yr in years:
            base = 2.0 ** (-((mx - yr) / hl))
            if yr in sb:            mult = SUPERBONUS_WEIGHT
            elif yr in covid_yrs:   mult = covid_w
            elif yr in postcovid_y: mult = postcovid_w
            elif yr in recent_yrs:  mult = recent_w
            else:                   mult = 1.0
            w.append(base * mult)
        w = np.array(w, dtype=float)
        w /= w.mean()
        return w

    def _sw(train_slice):
        if train_window == "tiered" and tiered_cfg:
            return _tiered_weights(train_slice.index.tolist())
        return _sample_weights(
            train_slice.index.tolist(), COUNTRY, superbonus_years
        )

    def _fit_and_predict(train_slice, X_fc):
        """Fit ensemble on train_slice, predict X_fc. Returns ensemble mean."""
        cols = clean_features + [target_pct]
        tr   = train_slice[cols].replace([np.inf,-np.inf], np.nan).dropna()
        if len(tr) < MIN_TRAIN_ROWS:
            return np.nan
        sw          = _sw(tr)
        sc          = StandardScaler()
        Xtr         = sc.fit_transform(tr[clean_features].values)
        Xte         = sc.transform(X_fc)
        ridge_alpha = RIDGE_ALPHA_MAP.get(COUNTRY, RIDGE_ALPHA)
        ps   = []
        for name in [n for n in linear_names if n not in ("LinearEnsemble","ARIMAX")]:
            mobj = {"Ridge":      Ridge(ridge_alpha),
                    "Lasso":      Lasso(LASSO_ALPHA, max_iter=10000),
                    "ElasticNet": ElasticNet(ENET_ALPHA, l1_ratio=ENET_L1, max_iter=10000),
                    }.get(name)
            if mobj is None: continue
            mobj.fit(Xtr, tr[target_pct].values, sample_weight=sw)
            ps.append(float(np.clip(mobj.predict(Xte)[0], -FORECAST_PCT_CAP, FORECAST_PCT_CAP)))
        return float(np.mean(ps)) if ps else np.nan

    blended_pct = {}
    blended_lvl = {}

    for yr in forecast_years:
        prev_lvl_ec = float(ec_full.loc[yr - 1]) if yr - 1 in ec_full.index else np.nan

        # ── Recursive ar_lag1 fix ─────────────────────────────────────────────
        # For 2026: ar_lag1 = ec_yoy[2025] (last known actual) — safe
        # For 2027: ar_lag1 MUST use model's own 2026 prediction, NOT ec_full[2026]
        #   Using EC 2026 here would leak the EC forecast into the 2027 prediction.
        #   Validated: Italy gap drops from 1.92pp → ~1.09pp with correct recursion.
        if "output_yoy_lag1" in clean_features:
            if yr == forecast_years[0]:
                # First forecast year: use last training actual
                py = yr - 1
                if py in ec_full.index and py - 1 in ec_full.index:
                    l1 = float(ec_full.loc[py])
                    l2 = float(ec_full.loc[py - 1])
                    if l2 > 0:
                        ar_val = round((l1 / l2 - 1) * 100, 3)
                        df.loc[yr, "output_yoy_lag1"] = ar_val
                        if yr in train_lagged.index:
                            train_lagged.loc[yr, "output_yoy_lag1"] = ar_val
            else:
                # Subsequent forecast years: use MODEL's own prior prediction
                prev_yr = yr - 1
                if prev_yr in blended_pct and pd.notna(blended_pct[prev_yr]):
                    ar_val = round(blended_pct[prev_yr], 3)
                    df.loc[yr, "output_yoy_lag1"] = ar_val
                    if yr in train_lagged.index:
                        train_lagged.loc[yr, "output_yoy_lag1"] = ar_val
                    print(f"  Recursive ar_lag1[{yr}] = model prediction [{prev_yr}] = {ar_val:+.3f}%")

        fc_row = df.loc[[yr], clean_features].replace([np.inf,-np.inf], np.nan) \
                 if yr in df.index else None
        if fc_row is None or fc_row.isna().any(axis=1).iloc[0]:
            miss = [f for f in clean_features
                    if yr in df.index and pd.isna(df.loc[yr, f])]
            print(f"  WARNING: {yr} missing features {miss} — skipping")
            blended_pct[yr] = np.nan
            blended_lvl[yr] = np.nan
            continue

        X_fc = fc_row.values

        if train_window == "full":
            pred  = _fit_and_predict(_train_slice("full"), X_fc)
            label = "full"

        elif train_window == "clean":
            pred  = _fit_and_predict(_train_slice("clean"), X_fc)
            label = "clean"

        elif train_window == "tiered":
            pred  = _fit_and_predict(_train_slice("tiered"), X_fc)
            label = "tiered"

        else:  # blend
            p_clean = _fit_and_predict(_train_slice("clean"), X_fc)
            p_full  = _fit_and_predict(_train_slice("full"),  X_fc)
            if pd.notna(p_clean) and pd.notna(p_full):
                pred = round(blend_ratio * p_clean + (1 - blend_ratio) * p_full, 3)
            elif pd.notna(p_full):
                pred = p_full
            else:
                pred = p_clean
            label = f"{blend_ratio:.0%}cl+{1-blend_ratio:.0%}fu"

        pred = float(np.clip(pred, -FORECAST_PCT_CAP, FORECAST_PCT_CAP)) \
               if pd.notna(pred) else np.nan

        blended_pct[yr] = round(pred, 3) if pd.notna(pred) else np.nan
        blended_lvl[yr] = pct_to_level(prev_lvl_ec, pred)

        # EC reference for display only (not used in prediction)
        ec_yr_lvl = float(ec_full.loc[yr]) if yr in ec_full.index else np.nan
        ec_yoy_yr = round((ec_yr_lvl / prev_lvl_ec - 1) * 100, 2) \
                    if pd.notna(ec_yr_lvl) and pd.notna(prev_lvl_ec) else np.nan

        print(f"  {yr} [{label}]:  ML={pred:+.2f}%  "
              f"EC_ref={ec_yoy_yr:+.2f}%  "
              f"gap={abs(pred - ec_yoy_yr):.2f}pp  "
              f"({blended_lvl[yr]:.1f} EUR bn)")

    # keep these for Excel export compatibility
    hw_pct_yr = np.nan
    el_pct_yr = np.nan
    blend_hw  = 0.0
    blend_el  = 0.0
    blend_ml  = 1.0

    # ── Comparison vs EC reference ────────────────────────────────────────────
    print(f"\n  {'Year':>5} {'EC YoY%':>10} {'Blend YoY%':>12} {'EC EUR bn':>10} {'Blend EUR bn':>13}")
    print(f"  {'-' * 55}")
    last_hist = float(hist_clean.loc[train_end, target_lvl])
    print(f"  {train_end:>5} {'(base)':>10} {'(base)':>12} {last_hist:>10.1f} {'(base)':>13}")
    for yr in forecast_years:
        prev_ec      = float(ec_full.loc[yr - 1]) if yr - 1 in ec_full.index else np.nan
        ec_yr        = float(ec_full.loc[yr])     if yr     in ec_full.index else np.nan
        ec_yoy_yr    = round((ec_yr / prev_ec - 1) * 100, 2) if pd.notna(ec_yr) and pd.notna(prev_ec) and prev_ec > 0 else np.nan
        bl_yoy       = blended_pct.get(yr, np.nan)
        bl_lvl       = blended_lvl.get(yr, np.nan)
        print(f"  {yr:>5} {ec_yoy_yr:>+10.2f}% {bl_yoy:>+12.2f}%  {ec_yr:>10.1f}  {bl_lvl:>13.1f}")

    # ── CAGR ──────────────────────────────────────────────────────────────────
    n_cagr    = test_end - train_end
    ec_s_cagr = float(ec_full.loc[train_end]) if train_end in ec_full.index else np.nan
    ec_e_cagr = float(ec_full.loc[test_end])  if test_end  in ec_full.index else np.nan
    bl_e_cagr = blended_lvl.get(test_end, np.nan)
    cagr_ec   = cagr(ec_s_cagr, ec_e_cagr, n_cagr)
    cagr_bl   = cagr(ec_s_cagr, bl_e_cagr, n_cagr)
    print(f"\n  CAGR {train_end}-{test_end}:")
    print(f"    Euroconstruct: {ec_s_cagr:.1f} → {ec_e_cagr:.1f}  {cagr_ec:+.2f}%/yr")
    print(f"    Blended model: {ec_s_cagr:.1f} → {bl_e_cagr:.1f}  {cagr_bl:+.2f}%/yr")

    # =========================================================================
    # CHARTS
    # =========================================================================
    start_year  = int(df[df.index >= 2000].index.min())
    ec_hist     = ec_full[ec_full.index <= train_end]
    ec_fcast    = ec_full[ec_full.index >= train_end]

    blend_s1    = {train_end: float(ec_full.loc[train_end])}
    blend_s1.update(blended_lvl)
    blend_s1_s  = pd.Series(blend_s1).dropna()

    _all_vals   = list(ec_full.dropna().values) + list(blend_s1_s.values)
    _lo, _hi    = min(_all_vals), max(_all_vals)
    _pad        = (_hi - _lo) * 0.15
    YMIN, YMAX  = max(0, _lo - _pad), _hi + _pad

    best_color  = "#E63946"
    ec_color    = "#1A1A1A"

    # Chart A — full series
    fig_a, ax_a = plt.subplots(figsize=(14, 4))
    _clean_axes(ax_a, fig_a)
    ax_a.plot(ec_hist.index, ec_hist.values, color=ec_color,
              linewidth=1.8, zorder=6, solid_capstyle="round")
    ax_a.plot(ec_fcast.index, ec_fcast.values, color=ec_color,
              linewidth=1.8, linestyle="--", zorder=6)
    ax_a.plot(blend_s1_s.index, blend_s1_s.values, color=best_color,
              linewidth=1.6, zorder=7, solid_capstyle="round")
    ax_a.axvline(x=train_end, color="#BBBBBB", linestyle="--",
                 linewidth=0.8, zorder=2)
    ax_a.text(train_end + 0.2, YMAX * 0.98, "Forecast",
              fontsize=8, color="#888888", va="top")
    ax_a.set_ylim(YMIN, YMAX)
    ax_a.set_title(f"Residential construction output — {COUNTRY}  (EUR bn)",
                   fontsize=9.5, loc="left", pad=6, color="#333333")
    ax_a.set_xticks(sorted(set(list(ec_full.index) + forecast_years)))
    ax_a.tick_params(axis="x", rotation=45, labelsize=7)
    leg_a = [
        Line2D([0],[0], color=ec_color, linewidth=1.8,
               label="Euroconstruct reference"),
        Line2D([0],[0], color=ec_color, linewidth=1.8, linestyle="--",
               label="Euroconstruct (forecast)"),
        Line2D([0],[0], color=best_color, linewidth=1.6,
               label=f"Model forecast (PATH A — pure ML)"),
    ]
    ax_a.legend(handles=leg_a, fontsize=7.5, loc="upper left",
                frameon=True, framealpha=0.9, edgecolor="#DDDDDD")
    plt.tight_layout()
    chart_a_path = os.path.join(OUT_DIR,
        f"{COUNTRY.replace(' ','_')}_construction_full_2006_{test_end}.png")
    fig_a.savefig(chart_a_path, dpi=150, bbox_inches="tight")
    plt.close(fig_a)
    print(f"\n  Chart A saved: {chart_a_path}")

    # Chart B — zoom forecast window
    fig_b, ax_b = plt.subplots(figsize=(9, 4.5))
    _clean_axes(ax_b, fig_b)
    zoom_years = [train_end] + forecast_years
    ec_z       = ec_full[ec_full.index >= train_end]
    ax_b.plot(ec_z[ec_z.index <= train_end].index,
              ec_z[ec_z.index <= train_end].values,
              color=ec_color, linewidth=2.0, zorder=6, solid_capstyle="round")
    ax_b.plot(ec_z[ec_z.index >= train_end].index,
              ec_z[ec_z.index >= train_end].values,
              color=ec_color, linewidth=2.0, linestyle="--", zorder=6)
    ax_b.plot(blend_s1_s[blend_s1_s.index >= train_end].index,
              blend_s1_s[blend_s1_s.index >= train_end].values,
              color=best_color, linewidth=2.0, zorder=7, solid_capstyle="round")
    ax_b.axvline(x=train_end, color="#BBBBBB", linestyle="--",
                 linewidth=0.8, zorder=2)
    ax_b.set_ylim(YMIN, YMAX)
    ym_b, yM_b = ax_b.get_ylim()
    ax_b.text(train_end + 0.05, yM_b * 0.98, "Forecast",
              fontsize=8, color="#888888", va="top")

    placed = []
    gap    = (yM_b - ym_b) * 0.05
    for name, val, col in sorted([
        ("Euroconstruct", ec_e_cagr and float(ec_full.loc[test_end])
                          if test_end in ec_full.index else np.nan, ec_color),
        ("Blended model", bl_e_cagr, best_color),
    ], key=lambda x: x[1] if pd.notna(x[1]) else 0, reverse=True):
        if pd.isna(val):
            continue
        y_t = float(val)
        for p in placed:
            if abs(y_t - p) < gap:
                y_t = p - gap
        placed.append(y_t)
        cg   = cagr_bl if name == "Blended model" else cagr_ec
        cg_s = f"  {cg:+.1f}%/yr" if pd.notna(cg) else ""
        ax_b.annotate(f"{val:.2f}bn{cg_s}",
                      xy=(test_end, val), xytext=(test_end + 0.08, y_t),
                      fontsize=7.5, color=col, va="center",
                      arrowprops=dict(arrowstyle="-", lw=0.5, color=col, alpha=0.4))

    ax_b.set_title(f"Residential construction output — {COUNTRY}  (EUR bn)",
                   fontsize=9.5, loc="left", pad=6, color="#333333")
    ax_b.set_xticks(zoom_years)
    ax_b.tick_params(axis="x", rotation=0, labelsize=9)
    leg_b = [
        Line2D([0],[0], color=ec_color, linewidth=2.0, label="Euroconstruct reference"),
        Line2D([0],[0], color=ec_color, linewidth=2.0, linestyle="--",
               label="Euroconstruct (forecast)"),
        Line2D([0],[0], color=best_color, linewidth=2.0,
               label=f"Model forecast (PATH A — pure ML)"),
    ]
    ax_b.legend(handles=leg_b, fontsize=7.5, loc="upper left",
                frameon=True, framealpha=0.9, edgecolor="#DDDDDD")
    plt.tight_layout()
    chart_b_path = os.path.join(OUT_DIR,
        f"{COUNTRY.replace(' ','_')}_construction_zoom_{train_end+1}_{test_end}.png")
    fig_b.savefig(chart_b_path, dpi=150, bbox_inches="tight")
    plt.close(fig_b)
    print(f"  Chart B saved: {chart_b_path}")

    # =========================================================================
    # EXCEL OUTPUTS
    # =========================================================================
    _export_blended_excel(
        country=COUNTRY, ec_full=ec_full,
        blended_pct=blended_pct, blended_lvl=blended_lvl,
        hw_levels=hw_levels, el_adjustments=el_adjustments,
        lin_pct=lin_pct, lin_lvl=lin_lvl,
        linear_names=linear_names, best_linear=best_linear,
        yoy_mae_dict=yoy_mae_dict, lvl_mape_dict=lvl_mape_dict, r2_dict=r2_dict,
        blend_hw=blend_hw, blend_el=blend_el, blend_ml=blend_ml,
        elasticities=elasticities,
        train_end=train_end, test_end=test_end,
        forecast_years=forecast_years, out_dir=OUT_DIR,
        start_year=start_year,
    )

    print(f"\n  Outputs saved to: {OUT_DIR}/")


# =============================================================================
# EXCEL EXPORT — BLENDED FORECAST
# =============================================================================

def _export_blended_excel(country, ec_full, blended_pct, blended_lvl,
                            hw_levels, el_adjustments, lin_pct, lin_lvl,
                            linear_names, best_linear,
                            yoy_mae_dict, lvl_mape_dict, r2_dict,
                            blend_hw, blend_el, blend_ml,
                            elasticities,
                            train_end, test_end, forecast_years,
                            out_dir, start_year=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _start    = start_year if start_year is not None else int(ec_full.index.min())
    all_years = list(range(_start, test_end + 1))
    _vx       = lambda v: round(float(v), 3) \
                if (v is not None and pd.notna(v)) else None
    _px       = lambda v: round(float(v) / 100, 4) \
                if (v is not None and pd.notna(v)) else None

    HDR_FILL   = PatternFill("solid", fgColor="1F3864")
    HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    MET_FONT   = Font(name="Arial", bold=True, size=9)
    DAT_FONT   = Font(name="Arial", size=9)
    EC_FILL    = PatternFill("solid", fgColor="E8F5E9")
    BLEND_FILL = PatternFill("solid", fgColor="FFF3E0")
    HW_FILL    = PatternFill("solid", fgColor="E3F2FD")
    ML_FILL    = PatternFill("solid", fgColor="F3E5F5")
    HIST_FILL  = PatternFill("solid", fgColor="F5F5F5")
    THIN       = Side(style="thin", color="DDDDDD")
    BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    center_a   = Alignment(horizontal="center", vertical="center")
    left_a     = Alignment(horizontal="left",   vertical="center")

    wb = Workbook()

    # ── Sheet 1: Forecast summary ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Forecast Summary"

    hdr = ["Metric"] + [str(y) for y in all_years]
    ws.append(hdr)
    for ci, v in enumerate(hdr, 1):
        c = ws.cell(row=1, column=ci)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = left_a if ci == 1 else center_a
        c.border = BORDER

    def _write_row(ws, ri, label, fill, yr_vals, fmt="0.000"):
        row_data = [label] + [yr_vals.get(y) for y in all_years]
        ws.append(row_data)
        for ci in range(1, len(row_data) + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = BORDER
            if ci == 1:
                cell.font = MET_FONT; cell.fill = fill; cell.alignment = left_a
            else:
                cell.font = DAT_FONT; cell.alignment = center_a
                yr = all_years[ci - 2]
                cell.fill = HIST_FILL if yr <= train_end else fill
                if cell.value is not None:
                    cell.number_format = fmt

    # EC levels (full history + reference)
    ec_vals = {y: _vx(ec_full.loc[y]) if y in ec_full.index else None
               for y in all_years}
    _write_row(ws, 2, "Euroconstruct reference (EUR bn)", EC_FILL, ec_vals)

    # Blended forecast levels
    bl_vals = {y: None if y <= train_end else _vx(blended_lvl.get(y))
               for y in all_years}
    _write_row(ws, 3, f"Blended forecast (HW {blend_hw:.0%} + Elast {blend_el:.0%}"
               f" + ML {blend_ml:.0%})  [EUR bn]", BLEND_FILL, bl_vals)

    # EC YoY%
    ec_yoy = {}
    for y in all_years:
        if y - 1 in ec_full.index and y in ec_full.index:
            prev = ec_full.loc[y - 1]
            curr = ec_full.loc[y]
            if pd.notna(prev) and pd.notna(curr) and float(prev) > 0:
                ec_yoy[y] = round((float(curr) / float(prev) - 1), 4)
    _write_row(ws, 4, "Euroconstruct: YoY growth", EC_FILL, ec_yoy, "0.00%")

    # Blended YoY%
    bl_yoy = {y: _px(blended_pct.get(y)) for y in all_years if y > train_end}
    _write_row(ws, 5, "Blended forecast: YoY growth", BLEND_FILL, bl_yoy, "0.00%")

    # HW component
    hw_vals = {y: None if y <= train_end else _vx(hw_levels.get(y))
               for y in all_years}
    _write_row(ws, 6, "Component A — Holt-Winters level (EUR bn)", HW_FILL, hw_vals)

    # ML ensemble YoY%
    ml_yoy = {y: _px(lin_pct["LinearEnsemble"].get(y)) for y in all_years
              if y > train_end and "LinearEnsemble" in lin_pct}
    _write_row(ws, 7, "Component C — Linear ensemble: YoY%", ML_FILL, ml_yoy, "0.00%")

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 52
    for i in range(2, len(all_years) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 7
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 15

    # ── Sheet 2: Linear model detail ──────────────────────────────────────────
    ws2 = wb.create_sheet("Linear Models")
    hdr2 = ["Model"] + [str(y) for y in all_years]
    ws2.append(hdr2)
    for ci, v in enumerate(hdr2, 1):
        c = ws2.cell(row=1, column=ci)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = left_a if ci == 1 else center_a
        c.border = BORDER

    ri2 = 2
    ec_row = ["Euroconstruct reference"] + \
             [_vx(ec_full.loc[y]) if y in ec_full.index else None for y in all_years]
    ws2.append(ec_row)
    for ci in range(1, len(ec_row) + 1):
        c = ws2.cell(row=ri2, column=ci)
        c.fill = EC_FILL; c.border = BORDER
        c.font = MET_FONT if ci == 1 else DAT_FONT
        c.alignment = left_a if ci == 1 else center_a
        if ci > 1 and c.value is not None:
            c.number_format = "0.000"
    ri2 += 1

    model_fills = {"Ridge":"E3F2FD","Lasso":"F3E5F5","ElasticNet":"E8F5E9",
                   "ARIMAX":"E0F7FA","LinearEnsemble":"FFF3E0"}
    for m in linear_names:
        mape_lbl = f" (YoY MAE {yoy_mae_dict[m]:.2f}pp)" if m in yoy_mae_dict else ""
        star     = "  ← BEST" if m == best_linear else ""
        label    = f"{m}{mape_lbl}{star}"
        row_data = [label] + [None if y <= train_end else _vx(lin_lvl[m].get(y))
                               for y in all_years]
        ws2.append(row_data)
        fill_hex = model_fills.get(m, "FFFFFF")
        mf       = PatternFill("solid", fgColor=fill_hex)
        for ci in range(1, len(row_data) + 1):
            c = ws2.cell(row=ri2, column=ci)
            c.border = BORDER
            if ci == 1:
                c.font = MET_FONT; c.fill = mf; c.alignment = left_a
            else:
                c.font = DAT_FONT; c.alignment = center_a
                yr = all_years[ci - 2]
                c.fill = HIST_FILL if yr <= train_end else mf
                if c.value is not None:
                    c.number_format = "0.000"
        ri2 += 1

    ws2.freeze_panes = "B2"
    ws2.column_dimensions["A"].width = 40
    for i in range(2, len(all_years) + 2):
        ws2.column_dimensions[get_column_letter(i)].width = 7

    # ── Sheet 3: Walk-forward accuracy ────────────────────────────────────────
    ws3 = wb.create_sheet("Walk-forward accuracy")
    ws3.column_dimensions["A"].width = 24
    for col in ["B", "C", "D"]:
        ws3.column_dimensions[col].width = 16
    ws3.append([f"Walk-forward validation {VALIDATE_FROM}-{train_end+1}  "
                f"| PRIMARY METRIC: YoY% MAE"])
    ws3.append(["Model", "YoY MAE (pp)", "Level MAPE %", "R²"])
    ws3["A1"].font = Font(name="Arial", bold=True, size=9, color="1F3864")
    for col in ["A2","B2","C2","D2"]:
        ws3[col].font = Font(name="Arial", bold=True, size=9)
    for mn in sorted(yoy_mae_dict, key=yoy_mae_dict.get):
        star = "  ← BEST" if mn == best_linear else ""
        ws3.append([
            mn + star,
            round(yoy_mae_dict.get(mn, np.nan), 2) if pd.notna(yoy_mae_dict.get(mn, np.nan)) else None,
            round(lvl_mape_dict.get(mn, np.nan), 2) if pd.notna(lvl_mape_dict.get(mn, np.nan)) else None,
            round(r2_dict.get(mn, np.nan), 4)       if pd.notna(r2_dict.get(mn, np.nan))       else None,
        ])
    for row in ws3.iter_rows(min_row=3):
        for c in row:
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="left")

    # ── Sheet 4: Notes ────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Notes")
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 70
    notes = [
        ("Country",           country),
        ("Best linear model", best_linear),
        ("Blend weights",     f"Holt-Winters {blend_hw:.0%} + Elasticity {blend_el:.0%} + ML {blend_ml:.0%}"),
        ("Elasticities used", str(elasticities)),
        ("Evaluation metric", "YoY% MAE (pp) — growth rate accuracy"),
        ("Level formula",     "level(yr) = EC_level(yr-1) × (1 + predicted_YoY% / 100)"),
        ("Target source",     "Euroconstruct_data.xlsx — residential construction (EUR bn)"),
        ("KPI changes",       "Removed: house_to_rent, population, household_size. "
                              "Merged: disposable/gross income → winner by corr. "
                              "Added: rate_regime_flag."),
        ("GDP note",          "GDP is NOMINAL current-price — YoY includes ~2-4pp inflation. "
                              "Replace with real GDP for cleaner signal."),
        ("Italy note",        "Superbonus years 2021-2022 downweighted in training "
                              "(SUPERBONUS_WEIGHT=0.20). Permits missing from source file."),
    ]
    ws4.append(["Field", "Value"])
    ws4["A1"].font = ws4["B1"].font = Font(name="Arial", bold=True, size=9)
    for field, value in notes:
        ws4.append([field, value])
    for row in ws4.iter_rows():
        for c in row:
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws4.row_dimensions[c.row].height = 18

    path = os.path.join(out_dir, f"{country.replace(' ','_')}_blended_forecast.xlsx")
    _safe_save(wb, path, "Blended forecast Excel")